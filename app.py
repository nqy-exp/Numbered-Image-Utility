import os
import sqlite3
import json
import io
import sys
import argparse
import pandas as pd  
from werkzeug.utils import secure_filename 
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, session, jsonify, send_file
from openpyxl.utils import get_column_letter 
from sync_engine import start_sync_service 
from dotenv import load_dotenv

load_dotenv() 


app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or os.urandom(24)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECTS_ROOT = os.path.join(BASE_DIR, 'projects')
IMAGE_ROOT = ""
DATABASE = ""
THUMBNAIL_ROOT = ""


# --- 1. 多语言字典 (保持不变) ---
LANGUAGES = {
    'zh': {
        # --- 通用/导航 ---
        'title': '编号图片实用小工具',
        'btn_back_index': '返回主页',
        'btn_back_prefix': '← 返回分组',
        'btn_back_to_view': '← 返回',
        'btn_back_to_select': '← 返回选择',
        'btn_back_prev': '← 返回上一页',
        'lang_switcher_zh': '中文', # 备用逻辑支持
        'lang_switcher_en': 'English',

        # --- 搜索与列表展示 ---
        'search_placeholder': '输入关键字...',
        'dir_title': '📂 分类目录',
        'search_title': '🔍 搜索结果: "{query}"',
        'empty_msg': '暂无任何记录',
        'no_results': '未找到匹配的文件',
        'count_files': '{} 张图片',
        'quick_jump': '快捷跳转到分组:',
        'filename_header': '文件名',
        'action_label': '操作',

        # --- 分组/目录视图 ---
        'sort_label': '排序方式:',
        'sort_default': '默认 (原始)',
        'sort_alpha_asc': '字母正序 (A-Z)',
        'sort_alpha_desc': '字母倒序 (Z-A)',
        'sort_date_asc': '日期正序 (旧-新)',
        'sort_date_desc': '日期倒序 (新-旧)',

        # --- 图片/卡片详情视图 ---
        'btn_view': '查看',
        'btn_edit': '编辑备注',
        'file_label': '文件名称:',
        'date_label': '📅 生成日期:',
        'exp_date_none': '未设置',
        'notes_label': '📝 备注说明:',
        'zoom_label': '🔍 缩放',
        'tags_input_hint': '🏷️ 标签 (输入 #标签 后按回车)',
        'btn_save': '💾 保存',
        'btn_close': '关闭页面',

        # --- 对比模式 ---
        'compare_mode_title': '⇄ 对比模式',
        'btn_compare': '对比',
        'btn_clear_all': '🗑️ 清空全部',
        'no_selection_msg': '暂未选择图片',
        'please_select_msg': '请选择图片！',
        'question_clear_all': '确定要清空所有选择的图片吗？此操作不可撤销！',
        'lock_unlocked': '🔓 锁定',
        'lock_locked': '🔒 已锁定',
        'btn_reorder': '🔄 重新排序',
        'modal_sort_title': '调整图片布局',
        'btn_add_row': '➕ 添加一行',
        'btn_remove_row': '➖ 减去一行',
        'btn_confirm': '确认保存',
        'btn_cancel': '取消',
        'msg_empty_slot': '空位',
        'text_tray': '待分配图片（拖拽到下方格子）',

        # --- 批量导入功能 ---
        'btn_import': '批量数据处理',
        'import_title': '批量数据处理',
        'import_subtitle': '通过上传 Excel 或 CSV 文件快速同步图片信息',
        'upload_prompt': '点击或将文件拖拽至此',
        'file_selected': '已选择: ',
        'step1_title': '空白模板',
        'step1_desc': '下载一个空文件，从头开始手动输入所有信息。',
        'btn_download_blank': '下载空白模板',
        'step2_title': '选择性填充 (智能)',
        'step2_desc': '勾选需要检查“空值”的列。我们将导出所有列，但仅针对那些未填写的行进行导出。',
        'btn_download_selective': '下载选择性模板',
        'col_check_notes': '备注',
        'col_check_date': '生成日期',
        'col_check_tags': '标签',
        'step3_title': '上传与同步',
        'step3_desc': '上传您填写好的 Excel/CSV 文件来批量更新数据库。',
        'step3_note': '⚠️ 注意：备注、日期与Tag为空即为删除原有信息。',
        'btn_start_import': '开始批量导入',
        'step4_title': '数据备份与导出',
        'step4_desc': '导出整个数据库，并支持自定义排序方式进行备份。',
        'sort_by_filename': '按文件名',
        'sort_by_date': '按生成日期',
        'sort_by_updated': '按最后更新时间',
        'order_asc': '升序（A-Z / 旧-新）',
        'order_desc': '降序（Z-A / 新-旧）',
        'btn_export_all': '📥 导出全部数据库数据',
        'guide_card_title': '💡 如何正确填写模板',
        'guide_filename': '文件名：必须与文件名完全匹配，包括大小写和完整后缀（T001-01.jpg）。',
        'guide_notes': '备注: 可以随意输入文字，系统会自动去除前后空格。',
        'guide_date': '生成日期: 请使用 YYYY-MM-DD 格式（例如 2026-08-06）。',
        'guide_tags': '标签 (tags): 使用逗号 ( , ) 分隔多个标签。',
        'guide_tags_hint': '提示：无需手动输入 "#" 号，系统会自动处理。全半角逗号会自动转换。',
        'tips_title': '💡 填写小贴士 (减少错误)',
        'tips_filename': '文件名：必须包含后缀 (如 T001-01.jpg)',
        'tips_date': '日期格式：必须是 YYYY-MM-DD',
        'tips_tags': '标签：用逗号分隔，无需加 # 号',
        'btn_export_errors': '📥 下载错误数据进行修改',
        'error_download_msg': '正在生成错误列表...',
        'error_modal_title': '⚠️ 导入出现问题',
        'btn_download_err': '📥 下载错误行进行修改',
        'err_filename_empty': '文件名不能为空',
        'err_file_not_found': '图片 "{name}" 在文件夹中不存在',
        'err_unsupported_format': '不支持的文件格式',
        'err_empty_file': '上传的文件是空的',
        'err_system_error': '系统错误: ',
        'err_date_format': '日期格式错误 (必须是 YYYY-MM-DD)',
        'err_extension_missing': '文件名缺少后缀名 (如 .jpg)',
        'err_invalid_tag': '标签格式不正确',
        'msg_success': '✅ 已成功保存',
        'msg_error': '❌ 保存失败',
        'msg_no_change': '⚠️ 没有检测到任何改动',
        'status_empty': '⚠️ 未填写',
        'user_guide': '使用说明',
        

    },
    'en': {
        # --- General / Navigation ---
        'title': 'Numbered Image Utility',
        'btn_back_index': 'Back to Home',
        'btn_back_prefix': '← Back to Prefix-Group',
        'btn_back_to_view': '← Back',
        'btn_back_to_select': '← Back to Select',
        'btn_back_prev': '← Go Back',
        'lang_switcher_zh': '中文',
        'lang_switcher_en': 'English',

        # --- Search & List Display ---
        'search_placeholder': 'Search...',
        'dir_title': '📂 Category Directory',
        'search_title': '🔍 Search Results: "{query}"',
        'empty_msg': 'No records found.',
        'no_results': 'No matching files found.',
        'count_files': '{} images',
        'quick_jump': 'Quick Jump to:',
        'filename_header': 'Filename',
        'action_label': 'Action',

        # --- Group / Directory View ---
        'sort_label': 'Sort by:',
        'sort_default': 'Default',
        'sort_alpha_asc': 'Alphabetical (A-Z)',
        'sort_alpha_desc': 'Alphabetical (Z-A)',
        'sort_date_asc': 'Date (Old - New)',
        'sort_date_desc': 'Date (New - Old)',

        # --- Image / Card Detail View ---
        'btn_view': 'View',
        'btn_edit': 'Edit Note',
        'file_label': 'File Name:',
        'date_label': '📅 Generated Date:',
        'exp_date_none': 'Not Set',
        'notes_label': '📝 Notes:',
        'zoom_label': '🔍 Zoom',
        'tags_input_hint': '🏷️ Tags (Type #tag then Enter)',
        'btn_save': '💾 Save',
        'btn_close': 'Close Page',

        # --- Compare Mode ---
        'compare_mode_title': '⇄ Comparison mode',
        'btn_compare': 'Compare',
        'btn_clear_all': '🗑️ Clear All',
        'no_selection_msg': 'No images selected yet.',
        'please_select_msg': 'Please select some images first!',
        'question_clear_all': 'Are you sure you want to clear all selected images? This action cannot be undone!',
        'lock_unlocked': '🔓 Lock',
        'lock_locked': '🔒 Locked',
        'btn_reorder': '🔄 Reorder',
        'modal_sort_title': 'Adjust Layout',
        'btn_add_row': '➕ Add Row',
        'btn_remove_row': '➖ Remove Row',
        'btn_confirm': 'Confirm',
        'btn_cancel': 'Cancel',
        'msg_empty_slot': 'Empty Slot',
        'text_tray': 'Images to be assigned (drag to the slots below)',

        # --- Batch Import Functionality ---
        'btn_import': 'Batch Data Processing',
        'import_title': 'Batch Data Processing',
        'import_subtitle': 'Upload Excel or CSV to sync image info quickly',
        'upload_prompt': 'Click or drag file here',
        'file_selected': 'Selected: ',
        'step1_title': 'Blank Template',
        'step1_desc': 'Download an empty file to manually enter all information from scratch.',
        'btn_download_blank': 'Download Blank Template',
        'step2_title': 'Selective Fill (Smart)',
        'step2_desc': 'Select columns to check for empty values. We will export all columns, but only for rows that need filling.',
        'btn_download_selective': 'Download Selective Template',
        'col_check_notes': 'Notes',
        'col_check_date': 'Generated Date',
        'col_check_tags': 'Tags',
        'step3_title': 'Upload & Sync',
        'step3_desc': 'Upload your filled Excel/CSV file to update the database.',
        'step3_note': '⚠️ Note: Leaving the notes, generated date, and tags empty will delete the existing information.',
        'btn_start_import': 'Start Batch Import',
        'step4_title': 'Data Backup & Export',
        'step4_desc': 'Export the entire database with custom sorting for backup purposes.',
        'sort_by_filename': 'Filename',
        'sort_by_date': 'Generated Date',
        'sort_by_updated': 'Last Updated',
        'order_asc': 'Ascending (A-Z / old-new)',
        'order_desc': 'Descending (Z-A / new-old)',
        'btn_export_all': '📥 Export All Database Data',
        'guide_card_title': '💡 Filling Guide',
        'guide_filename': 'Filename: Must match the filename exactly, include case and full extension (e.g., T001-01.jpg).',
        'guide_notes': 'Notes: Any text, spaces will be trimmed automatically.',
        'guide_date': 'Generated Date: Use YYYY-MM-DD format (e.g., 2026-08-06).',
        'guide_tags': 'Tags: Separate multiple tags with a comma (,).',
        'guide_tags_hint': 'Tip: No need to type "#", it is handled automatically.',
        'tips_title': '💡 Quick Tips (Avoid Errors)',
        'tips_filename': 'Filename: Must include extension (e.g. T001-01.jpg)',
        'tips_date': 'Date Format: Use YYYY-MM-DD',
        'tips_tags': 'Tags: Separate with commas, no # needed',
        'btn_export_errors': '📥 Download Error List to Fix',
        'error_download_msg': 'Generating error list...',
        'error_modal_title': '⚠️ Import Issues Detected',
        'btn_download_err': '📥 Download Error List',
        'err_filename_empty': 'Filename cannot be empty',
        'err_file_not_found': 'Image "{name}" not found in folder',
        'err_unsupported_format': 'Unsupported file format',
        'err_empty_file': 'The uploaded file is empty',
        'err_system_error': 'System error: ',
        'err_date_format': 'Invalid date format (Must be YYYY-MM-DD)',
        'err_extension_missing': 'File extension is missing (e.g., .jpg)',
        'err_invalid_tag': 'Invalid tag format',
        'msg_success': '✅ Saved Successfully',
        'msg_error': '❌ Save Failed',
        'msg_no_change': '⚠️ No changes detected',
        'status_empty': '⚠️ Empty',
        'user_guide': 'User Guide',
        

    }
}




# --- 2. 数据库操作函数 ---
def setup_project_paths(project_name):
    """根据项目名称，自动构建并创建该项目的完整路径结构"""
    global IMAGE_ROOT, DATABASE, THUMBNAIL_ROOT
    
    # 1. 构建该项目的根目录 (例如: .../projects/projecta)
    current_project_dir = os.path.join(PROJECTS_ROOT, project_name)
    
    # 2. 定义该项目内部的子文件夹路径
    IMAGE_ROOT = os.path.join(current_project_dir, 'images')
    THUMBNAIL_ROOT = os.path.join(current_project_dir, 'thumbnails')
    DATABASE = os.path.join(current_project_dir, 'experiments.db')

    # 3. 【核心】：自动创建文件夹，如果不存在的话
    os.makedirs(IMAGE_ROOT, exist_ok=True)
    os.makedirs(THUMBNAIL_ROOT, exist_ok=True)
    
    print(f"--- 项目加载成功 ---")
    print(f"项目名称: {project_name}")
    print(f"图片目录: {IMAGE_ROOT}")
    print(f"数据库路径: {DATABASE}")
    print(f"--------------------")

def get_db():
    conn = sqlite3.connect(DATABASE, timeout=20)
    conn.row_factory = sqlite3.Row
    return conn

def get_msg(key, lang, **kwargs):
    """
    根据 key 和语言获取对应的文字，并支持动态填充参数 (如 {name})
    """
    # 从全局 LANGUAGES 字典中获取对应的翻译
    # 如果找不到对应的 key，就返回 key 本身，防止程序崩溃
    msg = LANGUAGES.get(lang, LANGUAGES['en']).get(key, key)
    
    try:
        # 使用 kwargs 进行动态填充。例如：get_msg('err_file_not_found', 'zh', name='T01.jpg')
        # 会返回 "图片 'T01.jpg' 在文件夹中不存在"
        return msg.format(**kwargs)
    except KeyError:
        # 如果 format 里的参数没对上，就直接返回原文字
        return msg


def init_db():
    """初始化并强制升级数据库结构"""
    with get_db() as conn:
        # 1. 创建表（如果不存在）
        conn.execute('''
            CREATE TABLE IF NOT EXISTS image_notes (
                filename TEXT PRIMARY KEY,
                notes TEXT,
                experiment_date TEXT,
                tags TEXT,
                updated_at TEXT
            )
        ''')
        
        # 2. 检查并添加 tags 列（解决保存无效的问题）
        cursor = conn.execute("PRAGMA table_info(image_notes)")
        columns = [column[1] for column in cursor.fetchall()]
        if 'tags' not in columns:
            print("[系统提示] 检测到旧版数据库，正在强制升级 tags 字段...")
            conn.execute('ALTER TABLE image_notes ADD COLUMN tags TEXT')
            conn.commit()
            print("[系统提示] 数据库升级成功！")
        conn.commit()

# --- 1. 优化的全量数据获取函数 (解决 N+1 查询问题) ---
def get_all_files_with_data():
    """
    高度优化的数据获取函数。
    通过一次性批量查询数据库，将复杂度从 O(N) 降低到 O(1) 的查找效率。
    """
    if not os.path.exists(IMAGE_ROOT):
        return []

    # 获取文件系统中的所有图片文件名
    all_files = [f for f in os.listdir(IMAGE_ROOT) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))]
    if not all_files:
        return []

    conn = get_db()
    
    # 【核心改写】：使用 SQL 的 IN 子句一次性抓取所有相关记录
    placeholders = ', '.join(['?'] * len(all_files))
    query = f"SELECT * FROM image_notes WHERE filename IN ({placeholders})"
    
    # 将结果转换为一个以 filename 为 key 的字典，方便后续极速查找
    db_rows_map = {row['filename']: row for row in conn.execute(query, all_files).fetchall()}
    conn.close()

    combined_data = []
    for filename in all_files:
        row = db_rows_map.get(filename)
        
        parts = filename.split('-')
        prefix = "-".join(parts[:-1]) if len(parts) >= 2 else "未分类"
        display_name = os.path.splitext(filename)[0] 

        tags_list = []
        if row and row['tags']:
            tags_list = [t.strip() for t in row['tags'].split(',') if t.strip()]

        combined_data.append({
            'filename': filename,
            'display_name': display_name, 
            'prefix': prefix,
            'notes': row['notes'] if row else None,
            'experiment_date': row['experiment_date'] if row else None,
            'tags': tags_list,
            'updated_at': row['updated_at'] if row else None
        })

    return combined_data


# --- 2. 精准的混合搜索 API (解决搜索逻辑与性能问题) ---
@app.route('/api/search')
def api_search():
    """
    精准搜索 API：
    1. 搜索范围：仅限 文件名 (Filename) 和 标签 (Tags)。
    2. 排除内容：完全不搜索 备注 (Notes) 内容。
    3. 排序维度：支持按文件名、日期进行排序。
    """
    try:
        query = request.args.get('q', '').strip()
        sort = request.args.get('sort', 'default')

        if not query:
            return jsonify([])

        # 获取全量清单（包含物理文件和数据库记录的混合）
        all_items = get_all_files_with_data()

        # 解析搜索模式
        is_tag_mode = query.startswith('#')
        search_key = query[1:].lower() if is_tag_mode else query.lower()

        # 执行精准过滤
        if is_tag_mode:
            # 【标签模式】：只检查 tags 是否包含关键词
            filtered = [
                item for item in all_items 
                if item['tags'] and any(search_key in t.lower() for t in item['tags'])
            ]
        else:
            # 【文件名模式】：只检查 filename 是否包含关键词 (跳过 notes)
            filtered = [
                item for item in all_items 
                if search_key in item['filename'].lower()
            ]

        # 执行排序逻辑
        if sort == 'alpha_asc':
            filtered.sort(key=lambda x: x.get('prefix', ''))
        elif sort == 'alpha_desc':
            filtered.sort(key=lambda x: x.get('prefix', ''), reverse=True)
        elif sort == 'date_asc':
            filtered.sort(key=lambda x: x.get('experiment_date') or '')
        elif sort == 'date_desc':
            filtered.sort(key=lambda x: x.get('experiment_date') or '', reverse=True)

        return jsonify(filtered)

    except Exception as e:
        print(f"Search Error: {e}")
        return jsonify({"error": str(e)}), 500

# --- 3. 核心路由 ---

# --- 新增：把过滤逻辑抽离出来，方便两个地方调用 ---
def get_filtered_data(query='', sort='default'):
    all_items = get_all_files_with_data()
    filtered = all_items

    # 1. 过滤
    if query:
        is_tag_mode = query.startswith('#')
        search_key = query[1:].lower() if is_tag_mode else query.lower()
        if is_tag_mode:
            filtered = [i for i in all_items if i['tags'] and any(search_key in t.lower() for t in i['tags'])]
        else:
            filtered = [i for i in all_items if 
                        search_key in i['filename'].lower() or 
                        (i['notes'] and search_key in i['notes'].lower()) or
                        (i['tags'] and any(search_key in t.lower() for t in i['tags']))]

    # 2. 排序
    if sort == 'alpha_asc': filtered.sort(key=lambda x: x.get('prefix', ''))
    elif sort == 'alpha_desc': filtered.sort(key=lambda x: x.get('prefix', ''), reverse=True)
    elif sort == 'date_asc':
        # 按日期正序（旧到新），如果日期为空则排在最后
        filtered.sort(key=lambda x: x.get('experiment_date') or '9999-12-31')
    elif sort == 'date_desc':
        # 按日期倒序（新到旧），如果日期为空则排在最后
        filtered.sort(key=lambda x: x.get('experiment_date') or '0001-01-01', reverse=True)

    return filtered


@app.route('/')
def index():
    lang = session.get('lang', 'zh')
    query = request.args.get('q', '')
    sort = request.args.get('sort', 'default')
    
    # 1. 获取全量数据 (用于联想)
    full_library = get_all_files_with_data()
    
    # 2. 获取当前显示的过滤后数据 (用于页面渲染)
    display_items = get_filtered_data(query, sort)

    return render_template(
        'index.html', 
        all_items_json=json.dumps(full_library, ensure_ascii=False), # 当前要显示的
        library_json=json.dumps(full_library, ensure_ascii=False),     # 全量数据
        current_lang=lang, 
        lang_dict=LANGUAGES[lang],
        has_query=(query != "")
    )




@app.route('/set_lang/<lang>')
def set_lang(lang):
    if lang in LANGUAGES:
        session['lang'] = lang
    return redirect(request.args.get('next', '/'))

@app.route('/prefix/<prefix>')
def prefix_detail(prefix):
    lang = session.get('lang', 'zh')
    all_items = get_all_files_with_data()
    items = [i for i in all_items if i['prefix'] == prefix]
    return render_template('prefix_list.html', prefix=prefix, items=items, lang_dict=LANGUAGES[lang], current_lang=lang)

@app.route('/search')
def search():
    # 这个路由现在会被 index.html 的 JS 逻辑接管，但保留作为兜底
    return redirect(url_for('index'))



# --- [核心逻辑]：统一的清洗函数 ---
def clean_item_data(filename, notes=None, exp_date=None, tags=None):
    """标准化清洗，强制处理日期格式"""
    clean_filename = str(filename).strip() if filename else ""
    
    # 处理备注
    clean_notes = str(notes).strip() if notes is not None and str(notes).lower() != 'nan' else None
    
    # --- 【核心修复：日期清洗】 ---
    clean_date = None
    if exp_date is not None and str(exp_date).lower() != 'nan' and str(exp_date).strip() != "":
        date_str = str(exp_date).strip()
        # 如果字符串包含时间部分 (例如 "2026-08-06 00:00:00")，只取前10位
        if len(date_str) > 10 and ' ' in date_str:
            date_str = date_str.split(' ')[0]
        # 如果是 "2026/08/06" 这种格式，统一转为 "-"
        date_str = date_str.replace('/', '-')
        clean_date = date_str
    # ----------------------------

    # 处理标签
    clean_tags = None
    if tags is not None and str(tags).lower() != 'nan':
        tag_str = str(tags).replace('，', ',')
        tag_list = [t.strip().lstrip('#') for t in tag_str.split(',') if t.strip()]
        if tag_list:
            clean_tags = ",".join(tag_list)

    return clean_filename, clean_notes, clean_date, clean_tags



@app.route('/add_note/<path:filename>', methods=['GET', 'POST'])
def add_note(filename):
    # 【核心安全修复】：强制只提取文件名，防止路径穿越攻击 (Path Traversal)
    safe_filename = os.path.basename(filename)

    lang = session.get('lang', 'zh')
    conn = get_db()
    if request.method == 'POST':
        # 获取前端传来的数据
        raw_notes = request.form.get('notes', '')
        raw_exp_date = request.form.get('experiment_date', '')
        raw_tags = request.form.get('tags', '') # 假设前端传来的是逗号分隔字符串

        # 使用清洗函数标准化数据 (使用加固后的 safe_filename)
        fname, notes, exp_date, tags = clean_item_data(safe_filename, raw_notes, raw_exp_date, raw_tags)

        updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 使用 ON CONFLICT 进行“覆盖式更新”
        conn.execute('''
            INSERT INTO image_notes (filename, notes, experiment_date, tags, updated_at) 
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(filename) DO UPDATE SET 
                notes=excluded.notes, 
                experiment_date=excluded.experiment_date, 
                tags=excluded.tags,
                updated_at=excluded.updated_at
        ''', (fname, notes, exp_date, tags, updated_at))
        conn.commit()
        conn.close()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
            return jsonify({"status": "success"}), 200
        # 使用加固后的 safe_filename 进行重定向
        return redirect(url_for('view', filename=safe_filename))

    # GET 请求部分：使用加固后的 safe_filename 查询
    row = conn.execute('SELECT * FROM image_notes WHERE filename = ?', (safe_filename,)).fetchone()
    conn.close()
    tags_list = []
    if row and row['tags']:
        tags_list = [t.strip() for t in row['tags'].split(',') if t.strip()]

    return render_template(
        'edit.html', 
        filename=safe_filename, # 使用加固后的 safe_filename
        notes=row['notes'] if row else "", 
        exp_date=row['experiment_date'] if row else "",
        tags=tags_list, 
        lang_dict=LANGUAGES[lang],
        current_lang=lang
    )



@app.route('/view/<path:filename>')
def view(filename):
    # 【核心安全修复】：强制只取文件名，防止路径穿越攻击
    safe_filename = os.path.basename(filename)
    lang = session.get('lang', 'zh')
    conn = get_db()
    # 1. 获取数据库中的行数据 (使用 safe_filename)
    row = conn.execute('SELECT * FROM image_notes WHERE filename = ?', (safe_filename,)).fetchone()
    conn.close()

    # --- 【核心逻辑：解析标签】 ---
    tags_list = []
    if row and row['tags']:
        # 将数据库存储的字符串 "棉,测试" 转换为列表 ["棉", "测试"]
        tags_list = [t.strip() for t in row['tags'].split(',') if t.strip()]

    # --- 【核心逻辑：获取所属前缀，用于返回上一级】 ---
    current_prefix = ""
    if row:
        parts = safe_filename.split('-') # 使用 safe_filename
        # 如果文件名里有 '-'，则取前面的部分作为 prefix
        current_prefix = "-".join(parts[:-1]) if len(parts) >= 2 else ""
    else:
        # 如果数据库没找到记录（比如刚添加的），尝试根据文件名猜一下
        parts = safe_filename.split('-') # 使用 safe_filename
        current_prefix = "-".join(parts[:-1]) if len(parts) >= 2 else ""

    # --- 【渲染页面】 ---
    return render_template('view.html', 
        filename=safe_filename, # 使用加固后的文件名
        notes=row['notes'] if row else None, 
        exp_date=row['experiment_date'] if row else None, 
        tags=tags_list,        # 传入解析好的标签列表
        prefix=current_prefix, # 传入当前文件的 prefix，供 HTML 判断跳转路径
        lang_dict=LANGUAGES[lang],
        current_lang=lang
    )


@app.route('/images/<path:filename>')
def custom_static(filename):
    # 【核心安全修复】：强制只取文件名
    safe_filename = os.path.basename(filename)
    response = send_from_directory(IMAGE_ROOT, safe_filename)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


# --- 新增：专门用于访问缩略图的路由 ---
@app.route('/thumbnails/<path:filename>')
def get_thumbnail(filename):
    """
    这个路由负责响应 compare_index 页面发出的请求。
    它会去 thumbnails 文件夹里找文件并返回给浏览器。
    """
    # 【核心安全修复】：强制只取文件名
    safe_filename = os.path.basename(filename)
    response = send_from_directory(THUMBNAIL_ROOT, safe_filename)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


# --- 新增：用于访问根目录下非图片文件的路由 (如 favicon) ---
@app.route('/assets/<path:filename>')
def serve_assets(filename):
    # 【核心安全修复】：强制只取文件名
    safe_filename = os.path.basename(filename)
    response = send_from_directory(BASE_DIR, safe_filename)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


# --- 新增：图片对比功能路由 ---

@app.route('/compare_index')
def compare_index():
    lang = session.get('lang', 'zh')
    all_items = get_all_files_with_data() # 获取所有文件数据用于前端搜索
    return render_template(
        'compare_index.html', 
        all_items_json=json.dumps(all_items, ensure_ascii=False), 
        current_lang=lang, 
        lang_dict=LANGUAGES[lang]
    )

@app.route('/compare_view')
def compare_view():
    lang = session.get('lang', 'zh')
    # 从 URL 中获取选中的文件名列表 (例如: ?files=1.jpg,2.jpg)
    files_param = request.args.get('files', '')
    if not files_param:
        return redirect(url_for('index'))

    filenames = files_param.split(',')
    conn = get_db()
    compare_data = []

    for fname in filenames:
        # 【核心安全修复】：强制只取文件名，防止路径穿越攻击 (Path Traversal)
        safe_fname = os.path.basename(fname)
        
        # 查询数据库中每个文件的详细信息 (使用 safe_fname)
        row = conn.execute('SELECT * FROM image_notes WHERE filename = ?', (safe_fname,)).fetchone()
        
        # 解析标签
        tags_list = []
        if row and row['tags']:
            tags_list = [t.strip() for t in row['tags'].split(',') if t.strip()]

        compare_data.append({
            'filename': safe_fname, # 使用加固后的文件名
            'notes': row['notes'] if row else None,
            'experiment_date': row['experiment_date'] if row else None,
            'tags': tags_list
        })
    conn.close()

    return render_template(
        'compare_view.html',
        compare_items=compare_data,
        current_lang=lang,
        lang_dict=LANGUAGES[lang]
    )


@app.route('/import')
def import_page():
    lang = session.get('lang', 'zh')
    return render_template(
        'import.html', 
        current_lang=lang, 
        lang_dict=LANGUAGES[lang]
    )


@app.route('/view_com/<path:filename>')
def view_com(filename):
    # 【核心安全修复】：强制只取文件名，防止路径穿越攻击 (Path Traversal)
    safe_filename = os.path.basename(filename)
    
    lang = session.get('lang', 'zh')
    conn = get_db()
    # 使用 safe_filename 进行查询
    row = conn.execute('SELECT * FROM image_notes WHERE filename = ?', (safe_filename,)).fetchone()
    conn.close()

    tags_list = []
    if row and row['tags']:
        tags_list = [t.strip() for t in row['tags'].split(',') if t.strip()]

    return render_template(
        'view_com.html',  # 使用你专门创建的精简模板
        filename=safe_filename, # 使用加固后的文件名
        notes=row['notes'] if row else None, 
        exp_date=row['experiment_date'] if row else None, 
        tags=tags_list, 
        lang_dict=LANGUAGES[lang],
        current_lang=lang
    )



# --- 新增：批量导入 API 接口 ---
@app.route('/api/export_errors', methods=['POST'])
def export_errors():
    # 【修改点 1】：从 URL 参数中获取语言，默认为 'zh'
    lang = request.args.get('lang', 'zh')
    data = request.json  
    if not data:
        return jsonify({"status": "error", "message": "No error data provided"}), 400

    # 将接收到的 JSON 数据转回 DataFrame
    df_errors = pd.DataFrame(data)
    if 'experiment_date' in df_errors.columns:
        # 1. 先转成字符串
        # 2. 如果包含空格（即带了时间），只取空格前的部分
        df_errors['experiment_date'] = (
            df_errors['experiment_date']
            .astype(str)
            .apply(lambda x: x.split(' ')[0] if ' ' in x else x)
        )
    # 定义标准的内部键名顺序
    expected_order = ['filename', 'notes', 'experiment_date', 'tags']
    # 只保留存在的列，且按顺序排
    cols_to_keep = [c for c in expected_order if c in df_errors.columns]
    df_errors = df_errors[cols_to_keep]

    # --- 【修改点 2】：核心修复 - 表头翻译 ---
    mapping = COLUMN_MAPPING.get(lang, COLUMN_MAPPING['zh'])
    rename_dict = {k: v for k, v in mapping.items() if k in df_errors.columns}
    df_errors = df_errors.rename(columns=rename_dict)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_errors.to_excel(writer, index=False, sheet_name='Errors_to-Fix')
        worksheet = writer.sheets['Errors_to-Fix']

    # 设置简单的样式
    for i in range(1, len(df_errors) + 2):
        worksheet.row_dimensions[i].height = 35

    output.seek(0)
    return send_file(
        output, 
        as_attachment=True, 
        download_name="errors_to_fix.xlsx", 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --- 修改后的上传 API ---
@app.route('/api/upload_excel', methods=['POST'])
def upload_excel():
    lang = session.get('lang', 'zh') 
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": get_msg('err_no_file_part', lang)}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": get_msg('err_no_selected_file', lang)}), 400

    # 【关键】建立 中文 -> 英文 的映射表
    ZH_TO_EN_MAP = {
        '文件名': 'filename',
        '备注': 'notes',
        '生成日期': 'experiment_date',
        '标签': 'tags'
    }

    df = None  
    try:
        if file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file, dtype=str) 
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(file, dtype=str)
        else:
            return jsonify({"status": "error", "message": get_msg('err_unsupported_format', lang)}), 400

        if df is None or df.empty:
            return jsonify({"status": "error", "message": get_msg('err_empty_file', lang)}), 400

        conn = get_db()
        success_count = 0
        error_log = [] 

        for idx, row in df.iterrows():
            row_number = idx + 2  
            row_errors = []      
            
            raw_fname = row.get('文件名')
            raw_notes = row.get('备注')
            raw_date = row.get('生成日期')
            raw_tags = row.get('标签')

            # --- 第一步：文件名校验 ---
            valid_fname_for_cleaning = "" 
            if pd.isna(raw_fname) or str(raw_fname).strip() == '':
                row_errors.append(get_msg('err_filename_empty', lang))
            else:
                # 【核心安全修复】：使用 os.path.basename 防止路径穿越攻击 (Path Traversal)
                # 这确保了即使 Excel 里写的是 "../../etc/passwd"，程序也只会把它当做 "passwd" 处理
                fname_candidate = os.path.basename(str(raw_fname).strip())
                valid_fname_for_cleaning = fname_candidate 

                if '.' not in fname_candidate or len(fname_candidate.split('.')[-1]) < 2:
                    row_errors.append(get_msg('err_extension_missing', lang))
                
                # 使用加固后的文件名构建路径进行物理存在校验
                file_path = os.path.join(IMAGE_ROOT, fname_candidate)
                if not os.path.exists(file_path) or not os.path.isfile(file_path):
                    row_errors.append(get_msg('err_file_not_found', lang, name=fname_candidate))

            # --- 第二步：清洗与内容校验 ---
            clean_fname, clean_notes, clean_date, clean_tags = clean_item_data(
                valid_fname_for_cleaning, 
                raw_notes, 
                raw_date, 
                raw_tags
            )

            if clean_date:
                try:
                    datetime.strptime(clean_date, '%Y-%m-%d')
                except ValueError:
                    row_errors.append(get_msg('err_date_format', lang))

            # --- 第三步：根据错误列表决定操作 ---
            if row_errors:
                row_data = {}
                for col_zh in df.columns:
                    val = row.get(col_zh)
                    target_key = ZH_TO_EN_MAP.get(col_zh, col_zh) 
                    if pd.isna(val):
                        row_data[target_key] = None
                    else:
                        clean_val = str(val).strip()
                        row_data[target_key] = "" if clean_val == "NaT" else clean_val

                error_log.append({
                    "row": row_number,
                    "msg": " | ".join(row_errors), 
                    "data": row_data  
                })
            else:
                # 无错误，执行数据库操作
                try:
                    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    conn.execute('''
                        INSERT INTO image_notes (filename, notes, experiment_date, tags, updated_at) 
                        VALUES (?, ?, ?, ?, ?)
                        ON CONFLICT(filename) DO UPDATE SET 
                            notes=excluded.notes, 
                            experiment_date=excluded.experiment_date, 
                            tags=excluded.tags,
                            updated_at=excluded.updated_at
                    ''', (clean_fname, clean_notes, clean_date, clean_tags, updated_at))
                    success_count += 1
                except Exception as db_e:
                    row_data = {}
                    for col_zh in df.columns:
                        val = row.get(col_zh)
                        target_key = ZH_TO_EN_MAP.get(col_zh, col_zh)
                        if pd.isna(val):
                            row_data[target_key] = None
                        else:
                            clean_val = str(val).strip()
                            row_data[target_key] = "" if clean_val == "NaT" else clean_val

                    error_log.append({
                        "row": row_number,
                        "msg": f"Database Error: {str(db_e)}",
                        "data": row_data
                    })

            conn.commit()
        conn.close()

        if len(error_log) == 0:
            return jsonify({"status": "success", "message": f"Successfully imported {success_count} records!"})
        elif success_count > 0:
            return jsonify({
                "status": "partial_success", 
                "message": f"Imported {success_count}, but {len(error_log)} rows failed.",
                "errors": error_log 
            })
        else:
            return jsonify({"status": "error", "message": "All rows failed to import.", "errors": error_log})

    except Exception as e:
        print(f"Critical Upload Error: {e}") 
        return jsonify({"status": "error", "message": f"{get_msg('err_system_error', lang)}{str(e)}"}), 500





# --- 辅助映射配置 (放在函数外面，方便统一修改) ---
# 定义三种场景下的原始列名顺序
COLS_TEMPLATE = ['filename', 'notes', 'experiment_date', 'tags']
COLS_SELECTIVE = ['filename', 'notes', 'experiment_date', 'tags']
COLS_EXPORT_ALL = ['filename', 'notes', 'experiment_date', 'tags', 'updated_at']

# 定义中英文映射关系 (注意：这里对应的是你要求的“生成日期”和“最新更新时间”)
COLUMN_MAPPING = {
    'zh': {
        'filename': '文件名',
        'notes': '备注',
        'experiment_date': '生成日期',
        'tags': '标签',
        'updated_at': '最新更新时间'
    },
    'en': {
        'filename': 'filename',
        'notes': 'notes',
        'experiment_date': 'Generation Date',
        'tags': 'tags',
        'updated_at': 'Last Updated Time'
    }
}

# --- 1. 空白模板下载 ---
@app.route('/download_template')
def download_template():
    """下载带有示例行的纯空白模板 (严格控制：文件名 | 备注 | 生成日期 | 标签)"""
    lang = session.get('lang', 'zh')
    
    # 1. 获取当前语言的表头文字
    headers = [COLUMN_MAPPING[lang][col] for col in COLS_TEMPLATE]
    
    # --- 【核心改动：准备示例数据】 ---
    # 我们先用英文构建一个“原始数据”字典，这样逻辑最清晰
    example_data_raw = {
        'filename': 'T001-01.jpg',
        'notes': 'Example note content', # 英文示例
        'experiment_date': '2026-08-06',
        'tags': 'tag1,tag2'
    }

    # 如果是中文环境，我们要把示例内容也“翻译”一下（可选，但更贴心）
    if lang == 'zh':
        example_data_raw['notes'] = '这里是备注示例' # 中文示例
        # 注意：filename 和 experiment_date 保持原样即可

    # 根据当前语言的表头，构建最终要写入 Excel 的一行数据
    # 我们通过 mapping 将原始英文键值对转换为用户看到的中文/英文值
    example_row = []
    for col_key in COLS_TEMPLATE:  # 按照标准的顺序遍历
        val = example_data_raw[col_key]
        example_row.append(val)

    # 创建 DataFrame，并在第一行插入这个示例数据
    df = pd.DataFrame([example_row], columns=headers)
    # --- 【改动结束】 ---

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template')
        worksheet = writer.sheets['Template']

        # 设置行高 (len(df) 现在是 1，所以会处理第 1 和 第 2 行)
        for i in range(1, len(df) + 2): 
            worksheet.row_dimensions[i].height = 35  

        # 设置列宽
        for idx, col in enumerate(headers):
            column_len = len(col) + 15
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = column_len

    output.seek(0)
    return send_file(
        output, 
        as_attachment=True, 
        download_name="blank_template.xlsx", 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# --- 2. 选择性导出 ---
@app.route('/download_selective')
def download_selective():
    """选择性导出 (严格控制：不含 updated_at)"""
    lang = session.get('lang', 'zh')

    # 1. 获取参数
    selected_check_cols = request.args.get('cols', '').split(',')
    sort_by = request.args.get('sort', 'experiment_date') 
    order = request.args.get('order', 'asc')             

    if not selected_check_cols or selected_check_cols == ['']:
        return jsonify({"status": "error", "message": "No columns selected"}), 400

    # 2. 获取并合并数据
    if not os.path.exists(IMAGE_ROOT):
        return jsonify({"status": "error", "message": "Image directory not found"}), 500
    
    all_files = [f for f in os.listdir(IMAGE_ROOT) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))]
    df_files = pd.DataFrame({'filename': all_files})

    conn = get_db()
    df_db = pd.read_sql_query("SELECT * FROM image_notes", conn)
    conn.close()

    df = pd.merge(df_files, df_db, on='filename', how='left')

    if df.empty:
        return jsonify({"status": "error", "message": "No files found in directory"}), 400

    # 3. 执行过滤逻辑
    mask = pd.Series([True] * len(df))
    for col in selected_check_cols:
        if col in df.columns:
            is_empty = df[col].isna() | (df[col].astype(str).str.strip() == '')
            mask = mask & is_empty
    filtered_df = df[mask].copy()

    # 4. 执行排序逻辑
    if not filtered_df.empty:
        try:
            ascending = (order == 'asc')
            if sort_by in filtered_df.columns:
                filtered_df = filtered_df.sort_values(by=sort_by, ascending=ascending)
            else:
                filtered_df = filtered_df.sort_values(by='filename', ascending=ascending)
        except Exception as e:
            print(f"Selective sort error: {e}")

    # 5. 【核心修改】：严格控制列顺序和翻译，且不包含 updated_at
    # 我们直接使用 COLS_SELECTIVE 来提取原始数据，确保不会多出列
    filtered_df = filtered_df[COLS_SELECTIVE].copy()
    
    # 应用语言映射
    mapping = COLUMN_MAPPING[lang]
    filtered_df = filtered_df.rename(columns=mapping)

    # 6. 写入 Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        filtered_df.to_excel(writer, index=False, sheet_name='SelectiveData')
        worksheet = writer.sheets['SelectiveData']

        # 设置行高
        for i in range(1, len(filtered_df) + 2): 
            worksheet.row_dimensions[i].height = 35  

        # 设置列宽
        for idx, col in enumerate(filtered_df.columns):
            max_val_len = filtered_df[col].astype(str).map(len).max()
            column_len = max(max_val_len, len(col)) + 5
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = column_len

    output.seek(0)
    return send_file(
        output, 
        as_attachment=True, 
        download_name="selective_data.xlsx", 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --- 3. 全量导出 ---
@app.route('/export_all')
def export_all():
    """全量导出 (包含：文件名 | 备注 | 生成日期 | 标签 | 最新更新时间)"""
    lang = session.get('lang', 'zh')
    sort_by = request.args.get('sort', 'filename')
    order = request.args.get('order', 'asc')

    if not os.path.exists(IMAGE_ROOT):
        return jsonify({"status": "error", "message": "Image directory not found"}), 500
    
    all_files = [f for f in os.listdir(IMAGE_ROOT) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))]
    df_files = pd.DataFrame({'filename': all_files})

    conn = get_db()
    df_db = pd.read_sql_query("SELECT * FROM image_notes", conn)
    conn.close()

    df = pd.merge(df_files, df_db, on='filename', how='left')

    if df.empty:
        return jsonify({"status": "error", "message": "No files found"}), 400

    # 执行排序
    try:
        ascending = (order == 'asc')
        if sort_by in df.columns:
            df = df.sort_values(by=sort_by, ascending=ascending)
        else:
            df = df.sort_values(by='filename', ascending=ascending)
    except Exception as e:
        print(f"Sort error: {e}")

    # 5. 【核心修改】：严格使用 COLS_EXPORT_ALL 来包含 updated_at
    df = df[COLS_EXPORT_ALL].copy()
    
    # 应用语言映射
    mapping = COLUMN_MAPPING[lang]
    df = df.rename(columns=mapping)

    # 6. 写入 Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='FullData')
        worksheet = writer.sheets['FullData']

        # 设置行高
        for i in range(1, len(df) + 2):
            worksheet.row_dimensions[i].height = 35

        # 设置列宽
        for idx, col in enumerate(df.columns):
            max_val_len = df[col].astype(str).map(len).max()
            column_len = max(max_val_len, len(col)) + 15
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = column_len

    output.seek(0)
    return send_file(
        output, 
        as_attachment=True, 
        download_name="full_database_export.xlsx", 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# 在 app.py 的 if __name__ == '__main__': 部分增加 port 参数
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--project', type=str, default='default')
    parser.add_argument('--port', type=int, default=5001) # 新增端口参数
    args = parser.parse_args()

    setup_project_paths(args.project)
    init_db()
    observer = start_sync_service(IMAGE_ROOT, THUMBNAIL_ROOT)

    try:
        # 使用命令行传入的 port，而不是写死的 5001
        app.run(debug=True, port=args.port) 
    finally:
        observer.stop()
        observer.join()



